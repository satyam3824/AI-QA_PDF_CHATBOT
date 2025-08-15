import streamlit as st
import pandas as pd
import fitz  # PyMuPDF for PDFs
import csv
import os
import re
import requests
from openpyxl import load_workbook
import ezodf
import google.generativeai as genai
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

model = genai.GenerativeModel("gemini-2.5-flash")

st.set_page_config(page_title="Q/A File Bot", layout="wide")

st.header("📄 File Q&A Bot", divider='blue')
st.markdown("Ask questions about your documents (PDFs, CSVs, XLSX, ODS).")

# Store chat history
if "history" not in st.session_state:
    st.session_state.history = []

# --- Sidebar for File Upload and Controls ---
with st.sidebar:
    st.header("Upload your file")
    uploaded_file = st.file_uploader(
        "Choose a file", 
        type=["csv", "pdf", "xlsx", "ods"]
    )
    if st.button("Clear Chat History"):
        st.session_state.history = []
        st.rerun()

# Your core logic for fetching URL content, which is already good
def fetch_url_content(text):
    """Find first URL in text and fetch its content."""
    urls = re.findall(r'(https?://[^\s]+)', text)
    if urls:
        try:
            resp = requests.get(urls[0], timeout=10)
            resp.raise_for_status()
            return f"\n\n[Content from {urls[0]}]:\n{resp.text}"
        except Exception as e:
            return f"\n\n[Failed to fetch {urls[0]}: {e}]"
    return ""

if uploaded_file:
    st.info("File uploaded successfully. Scroll down to ask a question.")
    file_text = ""

    # Your file processing logic (no changes here)
    if uploaded_file.name.endswith(".csv"):
        file_text = uploaded_file.read().decode("utf-8")
    elif uploaded_file.name.endswith(".pdf"):
        pdf_doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
        for page in pdf_doc:
            file_text += page.get_text()
    elif uploaded_file.name.endswith(".xlsx"):
        wb = load_workbook(uploaded_file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                file_text += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
    elif uploaded_file.name.endswith(".ods"):
        ods_doc = ezodf.opendoc(uploaded_file)
        for sheet in ods_doc.sheets:
            for row in sheet.rows():
                cells = [cell.plaintext() for cell in row if cell.plaintext()]
                if cells:
                    file_text += " ".join(cells) + "\n"

    # Append fetched content if a URL is inside file_text
    file_text += fetch_url_content(file_text)

    # Display the chat messages
    for chat in st.session_state.history:
        with st.chat_message("user"):
            st.write(chat["q"])
        with st.chat_message("assistant"):
            st.write(chat["a"])
            with st.expander("📌 Source"):
                st.text(chat["source"])
    
    # Text input for new questions
    question = st.chat_input("Ask a question about your file...")
    
    if question:
        # Append the new question to history immediately
        st.session_state.history.append({"q": question, "a": "", "source": ""})
        # Rerun to show the user's message
        st.rerun()

# If there's a question in the last chat history entry, generate a response
if st.session_state.history and st.session_state.history[-1]["a"] == "" and uploaded_file:
    last_q = st.session_state.history[-1]["q"]
    with st.spinner("Generating response..."):
        full_prompt = f"Answer the following question based on the given file content:\n\n{file_text}\n\nQuestion: {last_q}\nAnswer in detail."
        response = model.generate_content(full_prompt)
        answer = response.text
        # Update the last entry with the bot's response
        st.session_state.history[-1]["a"] = answer
        st.session_state.history[-1]["source"] = file_text[:1000] + "..."
        st.rerun()